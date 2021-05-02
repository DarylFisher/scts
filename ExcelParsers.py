from openpyxl import load_workbook
from EntityMagic import *

class ExcelRowParser(): 

    def __init__(self,fname):
        self.fileName=fname
        self.workbook = load_workbook(fname)
        self.sheet = self.workbook.active
        print("file=", self.fileName)


    def loadHeader(self):
        self.headers = []
        colCount = 0
        for value in self.sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            if value != '':
                print("col =", colCount)
                self.headers += value
                self.colCount = colCount + 1
    
    def printHeaders(self):
        print ("Book = ", self.fileName)
        print ("** Headers ")
        for h in self.headers:
            print ("**** ", h)

    def parseData(self):
        a = []
        for row in self.sheet.iter_rows(min_row=2,values_only=True):
            ie = ImportEntity()
            j=0
            for r in row:
                c = EntityColumn(self.headers[j],r)
                j=j+1
                
                ie.addColumn(c)
            a.append(ie)
                   
        return a

            




def loadsheet():
    x = ExcelRowParser("TestActivity1.xlsx")
    x.loadHeader()
    x.printHeaders()
    p = x.parseData()
    

loadsheet()




            


    